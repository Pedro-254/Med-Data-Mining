from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import openpyxl
import time
from tqdm import tqdm
import re

def remover_linhas_duplicadas(arquivo_entrada, arquivo_saida):
    # Carrega o arquivo do Excel em um DataFrame do pandas
    df = pd.read_excel(arquivo_entrada)

    # Remove linhas duplicadas com base em todas as colunas
    df_sem_duplicatas = df.drop_duplicates()

    # Salva o DataFrame resultante de volta em um novo arquivo Excel
    df_sem_duplicatas.to_excel(arquivo_saida, index=False)

navegador = webdriver.Chrome()

#___________PROCESSO DE LOGIN_________________
navegador.get("https://v65.medx.med.br/login_unificado/loginunificado.html")
navegador.implicitly_wait(5)
elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-secondary")
elemento.click()

pag1_elemento_email = navegador.find_element(By.ID, "emailAssinante")
pag1_elemento_senha = navegador.find_element(By.ID, "senhaAssinante")

pag1_elemento_email.send_keys(input("Usuario_1: "))
pag1_elemento_senha.send_keys(input("Senha_1: "))

elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-block.btn-primary")
elemento.click()

pag2_elemento_email = navegador.find_element(By.ID, "usuario")
pag2_elemento_senha = navegador.find_element(By.ID, "senhaUsuario")

pag2_elemento_email.send_keys(input("Usuario_2: "))
pag2_elemento_senha.send_keys(input("Senha_2: "))

#___________PAGINA DE CONTATOS_________________
elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-block.btn-primary")
elemento.click()


# wait = WebDriverWait(navegador, 10)
# elemento = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".btn.btn-primary.ng-binding")))
for i in range(10):
    try:
        elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-primary.ng-binding")
        elemento.click()
        break
    except:
        None
    time.sleep(0.25)

navegador.get("https://v65.medx.med.br/pages_front_Desk/contatos.html")

#__________RESGATAR LISTA DE NOMES______________
# Carregar o arquivo Excel existente
wb = openpyxl.load_workbook('./Resultados/lista_de_nomes.xlsx')

# Selecionar a planilha ativa (por padrão, a primeira planilha)
sheet = wb.active

# Especificar a letra da coluna desejada (por exemplo, 'A' para a coluna A)
letra_coluna = 'A'

# Obter todos os valores da coluna, excluindo a primeira linha
coluna = sheet[letra_coluna][1:]

# Criar uma lista com os valores da coluna
lista_de_nomes = [celula.value for celula in coluna if celula.value is not None]
Lista_de_nomes_ausentes = []

#__________CRIAR PLANILHA DE EXCEL_____________
excel = openpyxl.Workbook()
sheet = excel.active
# Adicionar dados à planilha
sheet['A1'] = 'Nome'
sheet['B1'] = 'Genero'
sheet['C1'] = 'Idade'
sheet['D1'] = 'Celular'
sheet['E1'] = 'Email'
sheet['F1'] = 'Endereço'
sheet['G1'] = 'Observações'

#___________PESQUISAR CONTATO_________________
for i in tqdm(range(len(lista_de_nomes))):
    nome = lista_de_nomes[i]
    barra_de_pesquisa = navegador.find_element(By.ID, "inputBuscaContatos")
    barra_de_pesquisa.clear()
    barra_de_pesquisa.send_keys(nome)

    elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-primary.float-right")
    elemento.click()
    
    time.sleep(1)
    #_____________Trava para buscas com resultados multiplos_________
    elementos_tabela = navegador.find_elements(By.XPATH, "//tbody[@role='rowgroup']/tr")
    quantidade_elementos = len(elementos_tabela)
        
    if quantidade_elementos > 1 or quantidade_elementos == 0:
        print("entrou aquio" , nome)
        Lista_de_nomes_ausentes.append(nome)
        continue

    #____________Esperar Paciente_____________
    try:
        for k in range(10):
            nome_na_tela = navegador.find_element(By.XPATH, "//tr[@role='row']/td[3]")
            str_nome_na_tela = nome_na_tela.text

            str1 = set(nome)
            str2 = set(str_nome_na_tela)
            intersection = len(str1.intersection(str2))
            union = len(str1.union(str2))
            similarity = intersection / union
            if(similarity >= 0.8):
                break
            time.sleep(0.5)
        if(k == 9):
            raise ValueError("Paciente não encontrado!")
    except:
        Lista_de_nomes_ausentes.append(nome)
        continue

    nome_na_tela.click()

    #_______________Esperar Ficha________________
    try:
        for k in range(10):
            nome_ficha = navegador.find_element(By.XPATH, "//li[@class='tituloFichaPaciente']/span[@class='ng-binding']")
            nome_tela = navegador.find_element(By.XPATH, "//tr[@role='row']/td[3]")
            str_nome_ficha = nome_ficha.text
            str_nome_tela = nome_tela.text

            if(str_nome_ficha.lower() == str_nome_tela.lower()):
                break
            time.sleep(0.5)
        if(k == 9):
            raise ValueError("Paciente não encontrado!")
    except:
        Lista_de_nomes_ausentes.append(nome)
        continue

    #____________EXTRAIR DADOS________________
    lista_dados = navegador.find_element(By.XPATH, "//ul[@style='float: left; list-style: none; padding: 22px;']")
    itens = lista_dados.find_elements(By.XPATH, "li[@class='tituloFichaPaciente ng-scope']")


    # Nome / Genero / Idade / Celular / Email / Endereço / Observações
    Lista_de_valores = ["","","","","","",""]
    Lista_de_valores[0] = str(str_nome_ficha)
    
    #_______Genero e Idade__________
    Genero_Idade = navegador.find_element(By.CSS_SELECTOR, ".tituloFichaPaciente.ng-binding:nth-child(3)")
    Genero_Idade = str(Genero_Idade.text).split(',')
    Lista_de_valores[1] = Genero_Idade[0].replace(" ", "")
    Lista_de_valores[2] = re.sub(r'[^0-9]', '', Genero_Idade[1].replace(" ", ""))
    

    for item in itens:
        try:
            titulo = item.find_element(By.XPATH, "span[@class='textosFicha']").text
            valor = item.find_element(By.XPATH, "span[@class='ng-binding']").text

            # IDENTIFICAR QUAL O DADO E SALVAR ELE NO ARQUIVO EXCEL
            if(titulo == "Celular:"):
                # print("Celular: " + valor)
                Lista_de_valores[3] = valor
            elif(titulo == "Email:"):
                # print("Email: " + valor)
                Lista_de_valores[4] = valor
            elif(titulo == "Endereço:"):
                # print("Endereço: " + valor)
                Lista_de_valores[5] = valor
            elif(titulo == "Observações:"):
                # print("Observações: " + valor)
                Lista_de_valores[6] = valor
        except:
            None

    # Inserir valores na primeira linha da planilha
    for col_num, valor in enumerate(Lista_de_valores, 1):
        sheet.cell(row=i+2, column=col_num, value=valor)
    excel.save('./Resultados/Dados.xlsx')


#______________Limpando linhas Vazias e Duplicadas___________

# Carregue o arquivo Excel
caminho_arquivo = './Resultados/Dados.xlsx'
df = pd.read_excel(caminho_arquivo)

# Remova linhas vazias
df = df.dropna(how='all')

# Salve o DataFrame de volta no arquivo Excel
df.to_excel(caminho_arquivo, index=False)

remover_linhas_duplicadas('./Resultados/Dados.xlsx','./Resultados/Dados.xlsx')

#______________PACIENTES COM ERRO___________
print("Exibindo pacientes com erros: ")
for i in tqdm(range(len(Lista_de_nomes_ausentes))):
    nome = Lista_de_nomes_ausentes[i]
    barra_de_pesquisa = navegador.find_element(By.ID, "inputBuscaContatos")
    barra_de_pesquisa.clear()
    barra_de_pesquisa.send_keys(nome)

    elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-primary.float-right")
    elemento.click()
    input("Enter para continuar")