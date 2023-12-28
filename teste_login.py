from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
from tqdm import tqdm

navegador = webdriver.Chrome()

#___________PROCESSO DE LOGIN_________________
navegador.get("https://v65.medx.med.br/login_unificado/loginunificado.html")
navegador.implicitly_wait(5)
elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-secondary")
elemento.click()

pag1_elemento_email = navegador.find_element(By.ID, "emailAssinante")
pag1_elemento_senha = navegador.find_element(By.ID, "senhaAssinante")

pag1_elemento_email.send_keys("espacolisbsbadm@gmail.com")
pag1_elemento_senha.send_keys("FKhx765@")

elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-block.btn-primary")
elemento.click()

pag2_elemento_email = navegador.find_element(By.ID, "usuario")
pag2_elemento_senha = navegador.find_element(By.ID, "senhaUsuario")

pag2_elemento_email.send_keys("Dra. Carolina")
pag2_elemento_senha.send_keys("dracarol2021")

#___________PAGINA DE CONTATOS_________________
elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-block.btn-primary")
elemento.click()

time.sleep(1)
wait = WebDriverWait(navegador, 10)
elemento = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".btn.btn-primary.ng-binding")))
elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-primary.ng-binding")
elemento.click()

navegador.get("https://v65.medx.med.br/pages_front_Desk/contatos.html")

#__________RESGATAR LISTA DE NOMES______________
# Carregar o arquivo Excel existente
wb = openpyxl.load_workbook('./Resultados/lista_de_nomes.xlsx')

# Selecionar a planilha ativa (por padrão, a primeira planilha)
sheet = wb.active

# Especificar a letra da coluna desejada (por exemplo, 'A' para a coluna A)
letra_coluna = 'A'

# Obter todos os valores da coluna, excluindo a primeira linha
coluna = sheet[letra_coluna][1:]

# Criar uma lista com os valores da coluna
lista_de_nomes = [celula.value for celula in coluna if celula.value is not None]
Lista_de_nomes_ausentes = []

#__________CRIAR PLANILHA DE EXCEL_____________
excel = openpyxl.Workbook()
sheet = excel.active
# Adicionar dados à planilha
sheet['A1'] = 'Nome'
sheet['B1'] = 'Celular'
sheet['C1'] = 'Email'
sheet['D1'] = 'Endereço'
sheet['E1'] = 'Observações'

#___________PESQUISAR CONTATO_________________
for i in tqdm(range(len(lista_de_nomes))):
    nome = lista_de_nomes[i]
    barra_de_pesquisa = navegador.find_element(By.ID, "inputBuscaContatos")
    barra_de_pesquisa.clear()
    barra_de_pesquisa.send_keys(nome)

    elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-primary.float-right")
    elemento.click()
    
    time.sleep(1)
    #_____________Trava para buscas com resultados multiplos_________
    elementos_tabela = navegador.find_elements(By.XPATH, "//tbody[@role='rowgroup']/tr")
    quantidade_elementos = len(elementos_tabela)
        
    if quantidade_elementos > 1:
        Lista_de_nomes_ausentes.append(nome)
        continue

    #____________Esperar Paciente_____________
    try:
        for k in range(10):
            nome_na_tela = navegador.find_element(By.XPATH, "//tr[@role='row']/td[3]")
            str_nome_na_tela = nome_na_tela.text

            str1 = set(nome)
            str2 = set(str_nome_na_tela)
            intersection = len(str1.intersection(str2))
            union = len(str1.union(str2))
            similarity = intersection / union
            if(similarity >= 0.8):
                break
            time.sleep(0.5)
        if(k == 9):
            raise ValueError("Paciente não encontrado!")
    except:
        Lista_de_nomes_ausentes.append(nome)
        continue

    nome_na_tela.click()

    #_______________Esperar Ficha________________
    try:
        for k in range(10):
            nome_ficha = navegador.find_element(By.XPATH, "//li[@class='tituloFichaPaciente']/span[@class='ng-binding']")
            nome_tela = navegador.find_element(By.XPATH, "//tr[@role='row']/td[3]")
            str_nome_ficha = nome_ficha.text
            str_nome_tela = nome_tela.text
            print(str_nome_tela + "/" + str_nome_ficha)

            if(str_nome_ficha.lower() == str_nome_tela.lower()):
                break
            time.sleep(0.5)
        if(k == 9):
            raise ValueError("Paciente não encontrado!")
    except:
        Lista_de_nomes_ausentes.append(nome)
        continue

    #____________EXTRAIR DADOS________________
    lista_dados = navegador.find_element(By.XPATH, "//ul[@style='float: left; list-style: none; padding: 22px;']")
    itens = lista_dados.find_elements(By.XPATH, "li[@class='tituloFichaPaciente ng-scope']")


    # Nome / Celular / Email / Endereço / Observações
    Lista_de_valores = ["","","","",""]
    Lista_de_valores[0] = nome

    
    # print("Paciente: " + nome)
    

    for item in itens:
        try:
            titulo = item.find_element(By.XPATH, "span[@class='textosFicha']").text
            valor = item.find_element(By.XPATH, "span[@class='ng-binding']").text

            # IDENTIFICAR QUAL O DADO E SALVAR ELE NO ARQUIVO EXCEL
            if(titulo == "Celular:"):
                # print("Celular: " + valor)
                Lista_de_valores[1] = valor
            elif(titulo == "Email:"):
                # print("Email: " + valor)
                Lista_de_valores[2] = valor
            elif(titulo == "Endereço:"):
                # print("Endereço: " + valor)
                Lista_de_valores[3] = valor
            elif(titulo == "Observações:"):
                # print("Observações: " + valor)
                Lista_de_valores[4] = valor
        except:
            None

    # Inserir valores na primeira linha da planilha
    for col_num, valor in enumerate(Lista_de_valores, 1):
        sheet.cell(row=i+2, column=col_num, value=valor)
    excel.save('./Resultados/Dados.xlsx')

print("Exibindo pacientes com erros: ")
for i in tqdm(range(len(Lista_de_nomes_ausentes))):
    nome = Lista_de_nomes_ausentes[i]
    barra_de_pesquisa = navegador.find_element(By.ID, "inputBuscaContatos")
    barra_de_pesquisa.clear()
    barra_de_pesquisa.send_keys(nome)

    elemento = navegador.find_element(By.CSS_SELECTOR, ".btn.btn-primary.float-right")
    elemento.click()
    input("Enter para continuar")